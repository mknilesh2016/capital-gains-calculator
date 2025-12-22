"""
Parsers for Indian broker transaction files.

This module provides parsers for Groww capital gains reports
for stocks and mutual funds, as well as Zerodha P&L reports.
"""

from abc import ABC, abstractmethod
from typing import Dict, Any, Optional

from ..models import IndianGains


# Check if openpyxl is available
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class BaseIndianParser(ABC):
    """Abstract base class for Indian broker parsers."""
    
    @abstractmethod
    def parse(self, filepath: str) -> IndianGains:
        """
        Parse the file and return capital gains data.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            IndianGains object with parsed data
        """
        pass
    
    def _check_openpyxl(self, filepath: str) -> bool:
        """Check if openpyxl is available."""
        if not OPENPYXL_AVAILABLE:
            print(f"  [WARN] openpyxl not installed, cannot read {filepath}")
            return False
        return True


class IndianStocksParser(BaseIndianParser):
    """
    Parser for Groww Stocks Capital Gains Report.
    
    Reads the Excel file and extracts:
    - Short Term P&L
    - Long Term P&L
    - Dividends
    - Various charges (STT, brokerage, etc.)
    - Individual transaction details
    
    Expected format:
    - Row with "Short Term P&L" -> value in next column
    - Row with "Long Term P&L" -> value in next column
    - Section headers: "Intraday trades", "Short Term trades", "Long Term trades"
    - Transaction rows with columns: Stock name, ISIN, Quantity, Buy Date, etc.
    """
    
    CHARGE_FIELDS = [
        'Exchange Transaction Charges', 'SEBI Charges', 'STT',
        'Stamp Duty', 'Brokerage', 'DP Charges', 'Total GST'
    ]
    
    def parse(self, filepath: str) -> IndianGains:
        """
        Parse Indian stocks capital gains report.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            IndianGains object with STCG, LTCG, transactions, and charges
        """
        result = IndianGains(source='Indian Stocks')
        
        if not self._check_openpyxl(filepath):
            return result
        
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            current_section = None
            
            for row in ws.iter_rows():
                first_cell = row[0].value
                
                # Parse summary values
                if first_cell == 'Short Term P&L':
                    result.stcg = float(row[1].value or 0)
                elif first_cell == 'Long Term P&L':
                    result.ltcg = float(row[1].value or 0)
                elif first_cell == 'Dividends':
                    result.dividends = float(row[1].value or 0)
                elif first_cell in self.CHARGE_FIELDS:
                    result.charges[first_cell] = float(row[1].value or 0)
                
                # Identify sections
                elif first_cell == 'Intraday trades':
                    current_section = 'Intraday'
                elif first_cell == 'Short Term trades':
                    current_section = 'Short Term'
                elif first_cell == 'Long Term trades':
                    current_section = 'Long Term'
                
                # Parse transaction rows
                elif first_cell and first_cell != 'Stock name' and current_section:
                    txn = self._parse_transaction_row(row, current_section)
                    if txn:
                        result.transactions.append(txn)
            
            wb.close()
            
            print(f"   [OK] Indian Stocks: STCG = Rs.{result.stcg:,.2f}, LTCG = Rs.{result.ltcg:,.2f}")
            print(f"      {len(result.transactions)} transactions loaded")
            
        except Exception as e:
            print(f"   [ERROR] Error reading {filepath}: {e}")
        
        return result
    
    def _parse_transaction_row(self, row, section: str) -> Optional[Dict[str, Any]]:
        """Parse a single transaction row."""
        try:
            if row[2].value and row[3].value:  # Has quantity and buy date
                return {
                    'section': section,
                    'stock_name': str(row[0].value or ''),
                    'isin': str(row[1].value or ''),
                    'quantity': float(row[2].value or 0),
                    'buy_date': str(row[3].value or ''),
                    'buy_price': float(row[4].value or 0),
                    'buy_value': float(row[5].value or 0),
                    'sell_date': str(row[6].value or ''),
                    'sell_price': float(row[7].value or 0),
                    'sell_value': float(row[8].value or 0),
                    'pnl': float(row[9].value or 0),
                    'remark': str(row[10].value or '') if len(row) > 10 else ''
                }
        except (ValueError, TypeError):
            pass
        return None


class IndianMutualFundsParser(BaseIndianParser):
    """
    Parser for Mutual Funds Capital Gains Report.
    
    Reads the Excel file and extracts:
    - Taxable Short Term gains
    - Taxable Long Term gains
    - Individual redemption transactions
    
    Expected format:
    - Summary section with "Asset Class / Category" header
    - "Equity" row with STCG in column 4, LTCG in column 5
    - Transaction section starting with "Scheme Name" header
    """
    
    def parse(self, filepath: str) -> IndianGains:
        """
        Parse mutual funds capital gains report.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            IndianGains object with STCG, LTCG, and transactions
        """
        result = IndianGains(source='Indian Mutual Funds')
        
        if not self._check_openpyxl(filepath):
            return result
        
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            in_summary_section = False
            in_data_section = False
            
            for row in ws.iter_rows():
                # Look for summary section header
                if row[2].value == 'Asset Class / Category':
                    in_summary_section = True
                    continue
                
                # Parse Equity row in summary section
                if in_summary_section and row[2].value == 'Equity' and not in_data_section:
                    try:
                        result.stcg = float(row[3].value or 0)  # Taxable Short Term
                        result.ltcg = float(row[4].value or 0)  # Taxable Long Term
                    except (ValueError, TypeError):
                        pass
                    in_summary_section = False
                    continue
                
                # Identify transaction header row
                if row[0].value == 'Scheme Name':
                    in_data_section = True
                    continue
                
                # Parse transaction rows
                if in_data_section and row[0].value:
                    txn = self._parse_transaction_row(row)
                    if txn:
                        result.transactions.append(txn)
            
            wb.close()
            
            print(f"   [OK] Indian MFs: STCG = Rs.{result.stcg:,.2f}, LTCG = Rs.{result.ltcg:,.2f}")
            print(f"      {len(result.transactions)} transactions loaded")
            
        except Exception as e:
            print(f"   [ERROR] Error reading {filepath}: {e}")
        
        return result
    
    def _parse_transaction_row(self, row) -> Optional[Dict[str, Any]]:
        """Parse a single transaction row."""
        try:
            stcg = float(row[12].value or 0) if row[12].value else 0
            ltcg = float(row[13].value or 0) if row[13].value else 0
            
            return {
                'scheme_name': str(row[0].value or ''),
                'scheme_code': str(row[1].value or ''),
                'category': str(row[2].value or ''),
                'folio': str(row[3].value or ''),
                'purchase_date': str(row[5].value or ''),
                'quantity': float(row[6].value or 0),
                'purchase_price': float(row[7].value or 0),
                'redeem_date': str(row[9].value or ''),
                'redeem_price': float(row[11].value or 0),
                'stcg': stcg,
                'ltcg': ltcg,
                'classification': 'LTCG' if ltcg != 0 else 'STCG'
            }
        except (ValueError, TypeError, IndexError):
            pass
        return None


class ZerodhaPnLParser(BaseIndianParser):
    """
    Parser for Zerodha Profit & Loss Report (Equity).
    
    Reads the Excel file and extracts:
    - Realized P&L from closed positions
    - Various charges (STT, brokerage, etc.)
    - Individual stock transactions
    
    Expected format:
    - Summary section with "Realized P&L" value
    - Charges section with individual charge breakdowns
    - Transaction data starting with "Symbol" header row
    - Columns: Symbol, ISIN, Quantity, Buy Value, Sell Value, Realized P&L, etc.
    
    Note: Zerodha P&L report shows realized P&L which is treated as STCG
    (short-term capital gains) since the report doesn't distinguish 
    between short-term and long-term trades. For accurate LTCG/STCG
    classification, additional holding period information would be needed.
    """
    
    # Charge field mappings (Zerodha field name -> internal name)
    CHARGE_MAPPINGS = {
        'Brokerage - Z': 'Brokerage',
        'Exchange Transaction Charges - Z': 'Exchange Transaction Charges',
        'Clearing Charges - Z': 'Clearing Charges',
        'Central GST - Z': 'Central GST',
        'State GST - Z': 'State GST',
        'Integrated GST - Z': 'Integrated GST',
        'Securities Transaction Tax - Z': 'STT',
        'SEBI Turnover Fees - Z': 'SEBI Charges',
        'Stamp Duty - Z': 'Stamp Duty',
        'IPFT': 'IPFT',
    }
    
    def parse(self, filepath: str) -> IndianGains:
        """
        Parse Zerodha P&L report.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            IndianGains object with realized P&L and charges
        """
        result = IndianGains(source='Zerodha Stocks')
        
        if not self._check_openpyxl(filepath):
            return result
        
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            in_charges_section = False
            in_data_section = False
            total_realized_pnl = 0.0
            
            for row in ws.iter_rows():
                # Get values from the row (columns B and C in Excel = indices 1 and 2)
                col_b = row[1].value if len(row) > 1 else None
                col_c = row[2].value if len(row) > 2 else None
                
                # Parse summary "Realized P&L" value
                if col_b == 'Realized P&L' and col_c is not None:
                    try:
                        total_realized_pnl = float(col_c)
                    except (ValueError, TypeError):
                        pass
                
                # Identify Charges section
                if col_b == 'Charges' and col_c is None:
                    in_charges_section = True
                    continue
                
                # Parse individual charges
                if in_charges_section and col_b in self.CHARGE_MAPPINGS:
                    try:
                        charge_value = float(col_c or 0)
                        charge_name = self.CHARGE_MAPPINGS[col_b]
                        result.charges[charge_name] = charge_value
                    except (ValueError, TypeError):
                        pass
                
                # Check for Account Head header to stay in charges section
                if col_b == 'Account Head':
                    continue
                    
                # End charges section when we hit empty rows or transaction header
                if in_charges_section and col_b == 'Symbol':
                    in_charges_section = False
                    in_data_section = True
                    continue
                
                # Parse transaction rows
                if in_data_section and col_b and col_b != 'Symbol':
                    txn = self._parse_transaction_row(row)
                    if txn:
                        result.transactions.append(txn)
            
            wb.close()
            
            # Zerodha reports realized P&L without LTCG/STCG distinction
            # We'll treat it all as STCG (conservative approach)
            # The report includes both gains and losses in realized P&L
            result.stcg = total_realized_pnl
            result.ltcg = 0.0
            
            print(f"   [OK] Zerodha Stocks: Realized P&L = Rs.{result.stcg:,.2f}")
            print(f"      {len(result.transactions)} transactions loaded")
            total_charges = sum(result.charges.values())
            if total_charges > 0:
                print(f"      Total charges: Rs.{total_charges:,.2f}")
            
        except Exception as e:
            print(f"   [ERROR] Error reading {filepath}: {e}")
        
        return result
    
    def _parse_transaction_row(self, row) -> Optional[Dict[str, Any]]:
        """Parse a single transaction row from Zerodha P&L report."""
        try:
            # Skip rows without valid symbol or ISIN
            symbol = row[1].value if len(row) > 1 else None
            isin = row[2].value if len(row) > 2 else None
            
            if not symbol or not isin:
                return None
            
            # Skip header rows and invalid entries
            if symbol == 'Symbol' or isin == 'ISIN':
                return None
            
            quantity = float(row[3].value or 0) if len(row) > 3 and row[3].value else 0
            buy_value = float(row[4].value or 0) if len(row) > 4 and row[4].value else 0
            sell_value = float(row[5].value or 0) if len(row) > 5 and row[5].value else 0
            realized_pnl = float(row[6].value or 0) if len(row) > 6 and row[6].value else 0
            realized_pnl_pct = float(row[7].value or 0) if len(row) > 7 and row[7].value else 0
            open_quantity = float(row[9].value or 0) if len(row) > 9 and row[9].value else 0
            open_value = float(row[11].value or 0) if len(row) > 11 and row[11].value else 0
            unrealized_pnl = float(row[12].value or 0) if len(row) > 12 and row[12].value else 0
            
            return {
                'symbol': str(symbol),
                'isin': str(isin),
                'quantity': quantity,
                'buy_value': buy_value,
                'sell_value': sell_value,
                'realized_pnl': realized_pnl,
                'realized_pnl_pct': realized_pnl_pct,
                'open_quantity': open_quantity,
                'open_value': open_value,
                'unrealized_pnl': unrealized_pnl,
                'section': 'Short Term'  # Conservative default
            }
        except (ValueError, TypeError, IndexError):
            pass
        return None

