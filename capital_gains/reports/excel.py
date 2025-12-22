"""
Excel report generation module.

This module provides the ExcelReporter class for generating
formatted Excel workbooks with capital gains data.
"""

from datetime import datetime
from typing import List, Dict, Any

from ..models import SaleTransaction, IndianGains, TaxData
from ..utils import get_advance_tax_quarter, ADVANCE_TAX_QUARTERS


# Check if openpyxl is available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReporter:
    """
    Reporter for generating Excel workbooks.
    
    Creates multi-sheet workbooks with:
    - Summary sheet
    - Foreign stocks transactions
    - Exchange rates
    - Quarterly breakdown
    - Indian mutual funds (if data provided)
    - Indian stocks (if data provided)
    - Tax calculation
    """
    
    def __init__(self):
        """Initialize reporter with styles."""
        if not OPENPYXL_AVAILABLE:
            return
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.ltcg_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.stcg_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        self.summary_font = Font(bold=True, size=12)
        self.summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    def export(
        self,
        filepath: str,
        transactions: List[SaleTransaction],
        exchange_rates: Dict[str, float] = None,
        indian_gains: List[IndianGains] = None,
        tax_data: TaxData = None
    ) -> bool:
        """
        Export data to Excel workbook.
        
        Args:
            filepath: Output file path
            transactions: List of sale transactions
            exchange_rates: Dictionary of date -> rate
            indian_gains: List of Indian gains data
            tax_data: Tax calculation data
            
        Returns:
            True if export successful, False otherwise
        """
        if not OPENPYXL_AVAILABLE:
            print("\n[WARN] openpyxl not installed. Run: pip install openpyxl")
            return False
        
        exchange_rates = exchange_rates or {}
        indian_gains = indian_gains or []
        
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, transactions, indian_gains)
        self._create_transactions_sheet(wb, transactions)
        self._create_exchange_rates_sheet(wb, exchange_rates)
        self._create_quarterly_sheet(wb, transactions, indian_gains)
        
        if indian_gains:
            self._create_indian_gains_sheets(wb, indian_gains)
        
        if tax_data:
            self._create_tax_sheet(wb, tax_data, indian_gains)
        
        # Save workbook
        wb.save(filepath)
        print(f"[OK] Excel exported to: {filepath}")
        return True
    
    def _create_summary_sheet(self, wb, transactions, indian_gains):
        """Create the summary sheet."""
        ws = wb.active
        ws.title = "Summary"
        
        # Categorize transactions
        long_term = [t for t in transactions if t.is_long_term]
        short_term = [t for t in transactions if not t.is_long_term]
        eac_txns = [t for t in transactions if t.source == "EAC"]
        individual_txns = [t for t in transactions if t.source == "Individual"]
        
        # Capital Gains Classification
        row = 1
        ws.cell(row=row, column=1, value="CAPITAL GAINS CLASSIFICATION").font = self.summary_font
        ws.cell(row=row, column=1).fill = self.summary_fill
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        for col, header in enumerate(["Category", "Transactions", "Shares", 
                                       "Capital Gain (USD)", "Capital Gain (INR)"], 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            ws.cell(row=row, column=col).border = self.thin_border
        
        # Long Term row
        row += 1
        ws.cell(row=row, column=1, value="Foreign Stocks LTCG (> 2 years)")
        ws.cell(row=row, column=2, value=len(long_term))
        ws.cell(row=row, column=3, value=sum(t.shares for t in long_term))
        ws.cell(row=row, column=4, value=sum(t.capital_gain_usd for t in long_term))
        ws.cell(row=row, column=5, value=sum(t.capital_gain_inr for t in long_term))
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = self.ltcg_fill
            ws.cell(row=row, column=col).border = self.thin_border
        
        # Short Term row
        row += 1
        ws.cell(row=row, column=1, value="Foreign Stocks STCG (≤ 2 years)")
        ws.cell(row=row, column=2, value=len(short_term))
        ws.cell(row=row, column=3, value=sum(t.shares for t in short_term))
        ws.cell(row=row, column=4, value=sum(t.capital_gain_usd for t in short_term))
        ws.cell(row=row, column=5, value=sum(t.capital_gain_inr for t in short_term))
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = self.stcg_fill
            ws.cell(row=row, column=col).border = self.thin_border
        
        # Total row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL FOREIGN STOCKS").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(transactions)).font = Font(bold=True)
        ws.cell(row=row, column=3, value=sum(t.shares for t in transactions)).font = Font(bold=True)
        ws.cell(row=row, column=4, value=sum(t.capital_gain_usd for t in transactions)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=sum(t.capital_gain_inr for t in transactions)).font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).number_format = '₹#,##0.00'
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = self.thin_border
        
        # Add Indian investments section
        self._add_indian_summary(ws, row + 2, indian_gains, long_term, short_term)
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
    
    def _add_indian_summary(self, ws, start_row, indian_gains, long_term, short_term):
        """Add Indian investments summary to the sheet."""
        row = start_row
        ws.cell(row=row, column=1, value="INDIAN INVESTMENTS").font = self.summary_font
        ws.cell(row=row, column=1).fill = self.summary_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        for col, header in enumerate(["Source", "LTCG (INR)", "STCG (INR)", "Total (INR)"], 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            ws.cell(row=row, column=col).border = self.thin_border
        
        # Add row for each Indian source
        indian_ltcg_total = 0.0
        indian_stcg_total = 0.0
        
        for gains in indian_gains:
            row += 1
            # Map source names to display names
            display_name = self._get_indian_source_display_name(gains.source)
            ws.cell(row=row, column=1, value=display_name)
            ws.cell(row=row, column=2, value=gains.ltcg)
            ws.cell(row=row, column=3, value=gains.stcg)
            ws.cell(row=row, column=4, value=gains.total)
            
            indian_ltcg_total += gains.ltcg
            indian_stcg_total += gains.stcg
            
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = '₹#,##0.00'
                # Color code based on gain/loss
                value = ws.cell(row=row, column=col).value
                if value < 0:
                    ws.cell(row=row, column=col).fill = self.loss_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Indian Total row
        if len(indian_gains) > 1:
            row += 1
            ws.cell(row=row, column=1, value="Total Indian Investments").font = Font(bold=True)
            ws.cell(row=row, column=2, value=indian_ltcg_total).font = Font(bold=True)
            ws.cell(row=row, column=3, value=indian_stcg_total).font = Font(bold=True)
            ws.cell(row=row, column=4, value=indian_ltcg_total + indian_stcg_total).font = Font(bold=True)
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = '₹#,##0.00'
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.thin_border
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                )
        
        # Grand Total (All Sources)
        row += 2
        schwab_ltcg = sum(t.capital_gain_inr for t in long_term)
        schwab_stcg = sum(t.capital_gain_inr for t in short_term)
        grand_ltcg = schwab_ltcg + indian_ltcg_total
        grand_stcg = schwab_stcg + indian_stcg_total
        
        ws.cell(row=row, column=1, value="GRAND TOTAL (ALL SOURCES)").font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=grand_ltcg).font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=grand_stcg).font = Font(bold=True, size=12)
        ws.cell(row=row, column=4, value=grand_ltcg + grand_stcg).font = Font(bold=True, size=12)
        for col in range(2, 5):
            ws.cell(row=row, column=col).number_format = '₹#,##0.00'
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = self.thin_border
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            )
    
    def _get_indian_source_display_name(self, source: str) -> str:
        """Get display name for Indian investment sources."""
        display_names = {
            'Indian Stocks': 'Groww Stocks',
            'Indian Mutual Funds': 'Groww Mutual Funds',
            'Zerodha Stocks': 'Zerodha Stocks',
        }
        return display_names.get(source, source)
    
    def _create_transactions_sheet(self, wb, transactions):
        """Create the transactions sheet."""
        ws = wb.create_sheet("Schwab Foreign Stocks")
        
        headers = [
            'S.No', 'Source', 'Sale Date', 'Acquisition Date', 'Type', 'Symbol', 'Grant ID',
            'Shares', 'Holding Days', 'Holding Period', 'Classification',
            'Sale Price (USD)', 'Acquisition Price (USD)',
            'Sale Rate (INR/USD)', 'Acquisition Rate (INR/USD)',
            'Sale Price (INR)', 'Acquisition Price (INR)',
            'Total Sale (INR)', 'Total Acquisition (INR)',
            'Fees & Comm (USD)', 'Fees & Comm (INR)',
            'Capital Gain (USD)', 'Capital Gain (INR)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        ws.freeze_panes = 'A2'
        
        # Sort and add data
        sorted_txns = sorted(transactions, key=lambda x: (x.sale_date, x.symbol))
        
        for row_idx, txn in enumerate(sorted_txns, 2):
            row_data = [
                row_idx - 1, txn.source, txn.sale_date, txn.acquisition_date,
                txn.get_type_label(), txn.symbol, txn.grant_id or '',
                txn.shares, txn.holding_period_days, txn.get_holding_period_str(),
                'Long Term' if txn.is_long_term else 'Short Term',
                txn.sale_price_usd, txn.acquisition_price_usd,
                txn.sale_exchange_rate, txn.acquisition_exchange_rate,
                txn.sale_price_inr, txn.acquisition_price_inr,
                txn.total_sale_inr, txn.total_acquisition_inr,
                txn.fees_and_commissions_usd, txn.fees_and_commissions_inr,
                txn.capital_gain_usd, txn.capital_gain_inr
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.thin_border
                
                # Number formats
                if col_idx in [12, 13, 20]:
                    cell.number_format = '$#,##0.0000'
                elif col_idx in [14, 15]:
                    cell.number_format = '#,##0.0000'
                elif col_idx in [16, 17, 18, 19, 21, 23]:
                    cell.number_format = '₹#,##0.00'
                elif col_idx in [3, 4]:
                    cell.number_format = 'DD-MMM-YYYY'
            
            # Color code
            if txn.capital_gain_inr < 0:
                fill = self.loss_fill
            elif txn.is_long_term:
                fill = self.ltcg_fill
            else:
                fill = self.stcg_fill
            
            ws.cell(row=row_idx, column=11).fill = fill
            ws.cell(row=row_idx, column=23).fill = fill
        
        # Totals row
        total_row = len(sorted_txns) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=sum(t.shares for t in transactions)).font = Font(bold=True)
        ws.cell(row=total_row, column=22, value=sum(t.capital_gain_usd for t in transactions)).font = Font(bold=True)
        ws.cell(row=total_row, column=23, value=sum(t.capital_gain_inr for t in transactions)).font = Font(bold=True)
        ws.cell(row=total_row, column=22).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=23).number_format = '₹#,##0.00'
        
        # Column widths
        widths = [6, 12, 14, 14, 8, 8, 10, 10, 12, 12, 14, 16, 18, 16, 18, 16, 18, 18, 20, 14, 16, 18, 20]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    def _create_exchange_rates_sheet(self, wb, exchange_rates):
        """Create exchange rates sheet."""
        ws = wb.create_sheet("Exchange Rates")
        
        ws['A1'] = "Date"
        ws['B1'] = "USD-INR Rate"
        ws['C1'] = "Source"
        for col in ['A', 'B', 'C']:
            ws[f'{col}1'].font = self.header_font
            ws[f'{col}1'].fill = self.header_fill
        
        row = 2
        for date_str in sorted(exchange_rates.keys()):
            ws.cell(row=row, column=1, value=datetime.strptime(date_str, '%Y-%m-%d'))
            ws.cell(row=row, column=1).number_format = 'DD-MMM-YYYY'
            ws.cell(row=row, column=2, value=exchange_rates[date_str])
            ws.cell(row=row, column=2).number_format = '#,##0.0000'
            ws.cell(row=row, column=3, value="SBI TT Buy Rate")
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
    
    def _create_quarterly_sheet(self, wb, transactions, indian_gains):
        """Create quarterly breakdown sheet."""
        ws = wb.create_sheet("Quarterly Breakdown")
        quarters = ADVANCE_TAX_QUARTERS
        
        # Calculate foreign data
        foreign_data = {q: {'ltcg': 0.0, 'stcg': 0.0} for q in quarters}
        for txn in transactions:
            quarter = get_advance_tax_quarter(txn.sale_date)
            if quarter in foreign_data:
                if txn.is_long_term:
                    foreign_data[quarter]['ltcg'] += txn.capital_gain_inr
                else:
                    foreign_data[quarter]['stcg'] += txn.capital_gain_inr
        
        # Title
        ws.cell(row=1, column=1, value="CAPITAL GAINS - QUARTERLY BREAKDOWN").font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        
        self._add_quarterly_table(ws, 3, "FOREIGN STOCKS (Schwab)", foreign_data, quarters)
    
    def _add_quarterly_table(self, ws, start_row, title, data, quarters):
        """Add a quarterly breakdown table."""
        row = start_row
        
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = self.summary_fill
        ws.merge_cells(f'A{row}:G{row}')
        row += 1
        
        # Headers
        ws.cell(row=row, column=1, value="Sl").font = self.header_font
        ws.cell(row=row, column=2, value="Type").font = self.header_font
        for i, q in enumerate(quarters, 3):
            ws.cell(row=row, column=i, value=q).font = self.header_font
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = self.header_fill
            ws.cell(row=row, column=col).border = self.thin_border
        row += 1
        
        # LTCG row
        ws.cell(row=row, column=1, value=1)
        ws.cell(row=row, column=2, value="Long Term Capital Gain (LTCG)")
        for i, q in enumerate(quarters, 3):
            ws.cell(row=row, column=i, value=data[q]['ltcg'])
            ws.cell(row=row, column=i).number_format = '₹#,##0.00'
            ws.cell(row=row, column=i).fill = self.ltcg_fill
            ws.cell(row=row, column=i).border = self.thin_border
        ws.cell(row=row, column=1).border = self.thin_border
        ws.cell(row=row, column=2).border = self.thin_border
        row += 1
        
        # STCG row
        ws.cell(row=row, column=1, value=2)
        ws.cell(row=row, column=2, value="Short Term Capital Gain (STCG)")
        for i, q in enumerate(quarters, 3):
            ws.cell(row=row, column=i, value=data[q]['stcg'])
            ws.cell(row=row, column=i).number_format = '₹#,##0.00'
            ws.cell(row=row, column=i).fill = self.stcg_fill
            ws.cell(row=row, column=i).border = self.thin_border
        ws.cell(row=row, column=1).border = self.thin_border
        ws.cell(row=row, column=2).border = self.thin_border
        row += 1
        
        # Total row
        ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
        for i, q in enumerate(quarters, 3):
            total = data[q]['ltcg'] + data[q]['stcg']
            ws.cell(row=row, column=i, value=total)
            ws.cell(row=row, column=i).number_format = '₹#,##0.00'
            ws.cell(row=row, column=i).font = Font(bold=True)
            ws.cell(row=row, column=i).border = self.thin_border
        ws.cell(row=row, column=1).border = self.thin_border
        ws.cell(row=row, column=2).border = self.thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 32
        for i in range(3, 8):
            ws.column_dimensions[get_column_letter(i)].width = 16
        
        return row + 2
    
    def _create_indian_gains_sheets(self, wb, indian_gains):
        """Create sheets for all Indian investment sources."""
        for gains in indian_gains:
            display_name = self._get_indian_source_display_name(gains.source)
            sheet_name = display_name[:31]  # Excel sheet names max 31 chars
            
            ws = wb.create_sheet(sheet_name)
            
            # Title
            title = f"{display_name.upper()} - CAPITAL GAINS"
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            # Summary section
            ws.cell(row=3, column=1, value="Category")
            ws.cell(row=3, column=2, value="Amount (INR)")
            ws.cell(row=3, column=1).font = Font(bold=True)
            ws.cell(row=3, column=2).font = Font(bold=True)
            ws.cell(row=3, column=1).fill = self.header_fill
            ws.cell(row=3, column=2).fill = self.header_fill
            ws.cell(row=3, column=1).font = self.header_font
            ws.cell(row=3, column=2).font = self.header_font
            
            # LTCG row
            ws.cell(row=4, column=1, value="Long Term Capital Gain (LTCG)")
            ws.cell(row=4, column=2, value=gains.ltcg)
            ws.cell(row=4, column=2).number_format = '₹#,##0.00'
            ws.cell(row=4, column=2).fill = self.ltcg_fill if gains.ltcg >= 0 else self.loss_fill
            ws.cell(row=4, column=1).border = self.thin_border
            ws.cell(row=4, column=2).border = self.thin_border
            
            # STCG row
            ws.cell(row=5, column=1, value="Short Term Capital Gain (STCG)")
            ws.cell(row=5, column=2, value=gains.stcg)
            ws.cell(row=5, column=2).number_format = '₹#,##0.00'
            ws.cell(row=5, column=2).fill = self.stcg_fill if gains.stcg >= 0 else self.loss_fill
            ws.cell(row=5, column=1).border = self.thin_border
            ws.cell(row=5, column=2).border = self.thin_border
            
            # Total row
            ws.cell(row=6, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=6, column=2, value=gains.total).font = Font(bold=True)
            ws.cell(row=6, column=2).number_format = '₹#,##0.00'
            ws.cell(row=6, column=1).border = self.thin_border
            ws.cell(row=6, column=2).border = self.thin_border
            total_fill = self.ltcg_fill if gains.total >= 0 else self.loss_fill
            ws.cell(row=6, column=2).fill = total_fill
            
            # Charges section (if available)
            if gains.charges:
                row = 8
                ws.cell(row=row, column=1, value="CHARGES BREAKDOWN").font = Font(bold=True, size=12)
                ws.cell(row=row, column=1).fill = self.summary_fill
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
                
                ws.cell(row=row, column=1, value="Charge Type").font = Font(bold=True)
                ws.cell(row=row, column=2, value="Amount (INR)").font = Font(bold=True)
                ws.cell(row=row, column=1).fill = self.header_fill
                ws.cell(row=row, column=2).fill = self.header_fill
                ws.cell(row=row, column=1).font = self.header_font
                ws.cell(row=row, column=2).font = self.header_font
                row += 1
                
                total_charges = 0.0
                for charge_name, charge_value in gains.charges.items():
                    if charge_value > 0:
                        ws.cell(row=row, column=1, value=charge_name)
                        ws.cell(row=row, column=2, value=charge_value)
                        ws.cell(row=row, column=2).number_format = '₹#,##0.00'
                        ws.cell(row=row, column=1).border = self.thin_border
                        ws.cell(row=row, column=2).border = self.thin_border
                        total_charges += charge_value
                        row += 1
                
                ws.cell(row=row, column=1, value="TOTAL CHARGES").font = Font(bold=True)
                ws.cell(row=row, column=2, value=total_charges).font = Font(bold=True)
                ws.cell(row=row, column=2).number_format = '₹#,##0.00'
                ws.cell(row=row, column=1).border = self.thin_border
                ws.cell(row=row, column=2).border = self.thin_border
            
            # Transactions section (if available)
            if gains.transactions:
                self._add_indian_transactions_table(ws, gains)
            
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
    
    def _add_indian_transactions_table(self, ws, gains):
        """Add transactions table for Indian investments."""
        # Find the starting row (after charges or summary)
        start_row = 8 if not gains.charges else 8 + len([c for c in gains.charges.values() if c > 0]) + 4
        
        ws.cell(row=start_row, column=1, value="TRANSACTIONS DETAIL").font = Font(bold=True, size=12)
        ws.cell(row=start_row, column=1).fill = self.summary_fill
        
        # Determine columns based on source type
        if 'Zerodha' in gains.source:
            headers = ['Symbol', 'ISIN', 'Quantity', 'Buy Value', 'Sell Value', 'Realized P&L', 'P&L %']
            start_row += 1
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
                ws.cell(row=start_row, column=col).font = self.header_font
                ws.cell(row=start_row, column=col).fill = self.header_fill
                ws.cell(row=start_row, column=col).border = self.thin_border
            
            for i, txn in enumerate(gains.transactions, start_row + 1):
                row_data = [
                    txn.get('symbol', ''),
                    txn.get('isin', ''),
                    txn.get('quantity', 0),
                    txn.get('buy_value', 0),
                    txn.get('sell_value', 0),
                    txn.get('realized_pnl', 0),
                    txn.get('realized_pnl_pct', 0),
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=i, column=col, value=value)
                    ws.cell(row=i, column=col).border = self.thin_border
                    if col in [4, 5, 6]:
                        ws.cell(row=i, column=col).number_format = '₹#,##0.00'
                    elif col == 7:
                        ws.cell(row=i, column=col).number_format = '0.00%'
                
                # Color code based on P&L
                pnl = txn.get('realized_pnl', 0)
                fill = self.ltcg_fill if pnl >= 0 else self.loss_fill
                ws.cell(row=i, column=6).fill = fill
            
            # Expand columns for transaction detail
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 12
        
        elif 'Mutual Funds' in gains.source:
            headers = ['Scheme Name', 'Category', 'Folio', 'Purchase Date', 'Redeem Date', 'STCG', 'LTCG']
            start_row += 1
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
                ws.cell(row=start_row, column=col).font = self.header_font
                ws.cell(row=start_row, column=col).fill = self.header_fill
                ws.cell(row=start_row, column=col).border = self.thin_border
            
            for i, txn in enumerate(gains.transactions, start_row + 1):
                row_data = [
                    txn.get('scheme_name', ''),
                    txn.get('category', ''),
                    txn.get('folio', ''),
                    txn.get('purchase_date', ''),
                    txn.get('redeem_date', ''),
                    txn.get('stcg', 0),
                    txn.get('ltcg', 0),
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=i, column=col, value=value)
                    ws.cell(row=i, column=col).border = self.thin_border
                    if col in [6, 7]:
                        ws.cell(row=i, column=col).number_format = '₹#,##0.00'
            
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 16
        
        else:  # Indian Stocks (Groww)
            headers = ['Stock Name', 'ISIN', 'Section', 'Buy Date', 'Sell Date', 'Quantity', 'P&L']
            start_row += 1
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)
                ws.cell(row=start_row, column=col).font = self.header_font
                ws.cell(row=start_row, column=col).fill = self.header_fill
                ws.cell(row=start_row, column=col).border = self.thin_border
            
            for i, txn in enumerate(gains.transactions, start_row + 1):
                row_data = [
                    txn.get('stock_name', ''),
                    txn.get('isin', ''),
                    txn.get('section', ''),
                    txn.get('buy_date', ''),
                    txn.get('sell_date', ''),
                    txn.get('quantity', 0),
                    txn.get('pnl', 0),
                ]
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=i, column=col, value=value)
                    ws.cell(row=i, column=col).border = self.thin_border
                    if col == 7:
                        ws.cell(row=i, column=col).number_format = '₹#,##0.00'
                        pnl = txn.get('pnl', 0)
                        fill = self.ltcg_fill if pnl >= 0 else self.loss_fill
                        ws.cell(row=i, column=col).fill = fill
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['G'].width = 16
    
    def _create_tax_sheet(self, wb, tax_data: TaxData, indian_gains):
        """Create tax calculation sheet."""
        ws = wb.create_sheet("Tax Calculation")
        
        ws.cell(row=1, column=1, value="TAX LIABILITY CALCULATION").font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        # Tax rates
        ws.cell(row=3, column=1, value="Tax Rates Applied").font = self.summary_font
        ws.cell(row=3, column=1).fill = self.summary_fill
        ws.merge_cells('A3:C3')
        
        rates = [
            ("Indian LTCG Rate (12.5% + 15% SC + 4% Cess)", "14.95%"),
            ("Foreign LTCG Rate (12.5% + 15% SC + 4% Cess)", "14.95%"),
            ("Indian STCG Rate (20% + 15% SC + 4% Cess)", "23.92%"),
            ("Foreign STCG Rate (30% + 25% SC + 4% Cess)", "39%"),
            ("LTCG Exemption (Sec 112A)", f"₹{tax_data.ltcg_rebate:,.0f}")
        ]
        
        for i, (desc, value) in enumerate(rates, 4):
            ws.cell(row=i, column=1, value=desc)
            ws.cell(row=i, column=2, value=value)
        
        # Step 1: LTCG Exemption
        row = 10
        ws.cell(row=row, column=1, value="Step 1: LTCG Exemption (Section 112A)").font = self.summary_font
        ws.cell(row=row, column=1).fill = self.summary_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        exemption_items = [
            ("Indian LTCG (before exemption)", tax_data.indian_ltcg),
            ("Less: LTCG Exemption Used", -tax_data.rebate_used),
            ("Indian LTCG (after exemption)", tax_data.indian_ltcg_after_rebate),
        ]
        
        for desc, value in exemption_items:
            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = '₹#,##0.00'
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=2).border = self.thin_border
            row += 1
        
        row += 1
        
        # Step 2: Loss Set-off
        ws.cell(row=row, column=1, value="Step 2: Loss Set-off").font = self.summary_font
        ws.cell(row=row, column=1).fill = self.summary_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        # Gains before set-off
        ws.cell(row=row, column=1, value="Gains Before Set-off:").font = Font(bold=True, italic=True)
        row += 1
        
        gains_items = [
            ("  Foreign LTCG (Schwab)", tax_data.foreign_ltcg_gain),
            ("  Indian LTCG (after exemption)", tax_data.indian_ltcg_gain),
            ("  Foreign STCG (Schwab)", tax_data.foreign_stcg_gain),
            ("  Indian STCG", tax_data.indian_stcg_gain),
        ]
        
        for desc, value in gains_items:
            ws.cell(row=row, column=1, value=desc)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=2).number_format = '₹#,##0.00'
            if value > 0:
                ws.cell(row=row, column=2).fill = self.ltcg_fill
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=2).border = self.thin_border
            row += 1
        
        row += 1
        
        # Losses before set-off
        total_ltcg_loss = tax_data.foreign_ltcg_loss + tax_data.indian_ltcg_loss
        total_stcg_loss = tax_data.foreign_stcg_loss + tax_data.indian_stcg_loss
        
        if total_ltcg_loss > 0 or total_stcg_loss > 0:
            ws.cell(row=row, column=1, value="Losses Before Set-off:").font = Font(bold=True, italic=True)
            row += 1
            
            losses_items = []
            if tax_data.foreign_ltcg_loss > 0:
                losses_items.append(("  Foreign LTCG Loss (Schwab)", -tax_data.foreign_ltcg_loss))
            if tax_data.indian_ltcg_loss > 0:
                losses_items.append(("  Indian LTCG Loss", -tax_data.indian_ltcg_loss))
            if tax_data.foreign_stcg_loss > 0:
                losses_items.append(("  Foreign STCG Loss (Schwab)", -tax_data.foreign_stcg_loss))
            if tax_data.indian_stcg_loss > 0:
                losses_items.append(("  Indian STCG Loss", -tax_data.indian_stcg_loss))
            
            for desc, value in losses_items:
                ws.cell(row=row, column=1, value=desc)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = '₹#,##0.00'
                ws.cell(row=row, column=2).fill = self.loss_fill
                ws.cell(row=row, column=1).border = self.thin_border
                ws.cell(row=row, column=2).border = self.thin_border
                row += 1
            
            row += 1
        
        # Set-offs applied
        has_setoffs = (tax_data.stcg_loss_vs_foreign_stcg > 0 or 
                       tax_data.stcg_loss_vs_indian_stcg > 0 or 
                       tax_data.stcg_loss_vs_ltcg > 0 or 
                       tax_data.ltcg_loss_vs_ltcg > 0)
        
        if has_setoffs:
            ws.cell(row=row, column=1, value="Set-offs Applied:").font = Font(bold=True, italic=True)
            row += 1
            
            setoff_items = []
            if tax_data.stcg_loss_vs_foreign_stcg > 0:
                setoff_items.append(("  STCG Loss → Foreign STCG Gain", -tax_data.stcg_loss_vs_foreign_stcg))
            if tax_data.stcg_loss_vs_indian_stcg > 0:
                setoff_items.append(("  STCG Loss → Indian STCG Gain", -tax_data.stcg_loss_vs_indian_stcg))
            if tax_data.stcg_loss_vs_ltcg > 0:
                setoff_items.append(("  STCG Loss → LTCG Gain", -tax_data.stcg_loss_vs_ltcg))
            if tax_data.ltcg_loss_vs_ltcg > 0:
                setoff_items.append(("  LTCG Loss → LTCG Gain", -tax_data.ltcg_loss_vs_ltcg))
            
            for desc, value in setoff_items:
                ws.cell(row=row, column=1, value=desc)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = '₹#,##0.00'
                ws.cell(row=row, column=2).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                ws.cell(row=row, column=1).border = self.thin_border
                ws.cell(row=row, column=2).border = self.thin_border
                row += 1
            
            row += 1
        
        # Net taxable amounts
        ws.cell(row=row, column=1, value="Net Taxable Amounts:").font = Font(bold=True, italic=True)
        row += 1
        
        net_items = [
            ("  NET LTCG (Taxable)", tax_data.net_ltcg),
            ("  NET STCG (Taxable)", tax_data.net_stcg),
        ]
        
        for desc, value in net_items:
            ws.cell(row=row, column=1, value=desc).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value).font = Font(bold=True)
            ws.cell(row=row, column=2).number_format = '₹#,##0.00'
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=2).border = self.thin_border
            row += 1
        
        row += 1
        
        # Step 3: Tax calculation
        ws.cell(row=row, column=1, value="Step 3: Tax Calculation").font = self.summary_font
        ws.cell(row=row, column=1).fill = self.summary_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        calc_items = [
            ("Foreign LTCG Tax @ 14.95%", tax_data.foreign_ltcg_tax),
            ("Indian LTCG Tax @ 14.95%", tax_data.indian_ltcg_tax),
            ("Total LTCG Tax", tax_data.ltcg_tax),
            ("", ""),
            ("Indian STCG Tax @ 23.92%", tax_data.indian_stcg_tax),
            ("Foreign STCG Tax @ 39%", tax_data.foreign_stcg_tax),
            ("Total STCG Tax", tax_data.stcg_tax),
            ("", ""),
            ("TOTAL TAX", tax_data.total_tax),
            ("Less: Taxes Paid", tax_data.taxes_paid),
            ("", ""),
        ]
        
        for desc, value in calc_items:
            ws.cell(row=row, column=1, value=desc)
            if value != "":
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = '₹#,##0.00'
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=2).border = self.thin_border
            row += 1
        
        # Final liability
        label = "TAX PAYABLE" if tax_data.tax_liability > 0 else "TAX REFUND DUE"
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=abs(tax_data.tax_liability)).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).number_format = '₹#,##0.00'
        fill = self.loss_fill if tax_data.tax_liability > 0 else self.ltcg_fill
        ws.cell(row=row, column=2).fill = fill
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

