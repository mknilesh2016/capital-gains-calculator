"""
Unit tests for transaction parsers.
"""

import pytest
import os
import tempfile
from datetime import datetime

from capital_gains.parsers.schwab import SchwabEACParser, SchwabIndividualParser
from capital_gains.parsers.indian import ZerodhaPnLParser

# Check if openpyxl is available for Zerodha tests
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TestSchwabEACParser:
    """Tests for SchwabEACParser class."""
    
    @pytest.fixture
    def parser(self):
        """Create parser instance."""
        return SchwabEACParser()
    
    def test_parse_rsu_sale(self, parser):
        """Test parsing RSU sale transaction."""
        transactions = [
            {
                "Action": "Sale",
                "Date": "04/15/2025",
                "Symbol": "AAPL",
                "FeesAndCommissions": "$10.00",
                "TransactionDetails": [
                    {
                        "Details": {
                            "Type": "RS",
                            "Shares": "100",
                            "SalePrice": "$150.00",
                            "GrossProceeds": "$15000.00",
                            "VestDate": "01/15/2023",
                            "VestFairMarketValue": "$120.00",
                            "GrantId": "G12345"
                        }
                    }
                ]
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 1
        txn = result[0]
        assert txn.symbol == "AAPL"
        assert txn.shares == 100
        assert txn.sale_price_usd == 150.0
        assert txn.acquisition_price_usd == 120.0
        assert txn.stock_type == "RS"
        assert txn.source == "EAC"
        assert txn.grant_id == "G12345"
        assert txn.is_long_term  # >2 years
    
    def test_parse_espp_sale(self, parser):
        """Test parsing ESPP sale transaction."""
        transactions = [
            {
                "Action": "Sale",
                "Date": "04/15/2025",
                "Symbol": "AAPL",
                "FeesAndCommissions": "$5.00",
                "TransactionDetails": [
                    {
                        "Details": {
                            "Type": "ESPP",
                            "Shares": "50",
                            "SalePrice": "$150.00",
                            "GrossProceeds": "$7500.00",
                            "PurchaseDate": "01/15/2024",
                            "PurchaseFairMarketValue": "$130.00"
                        }
                    }
                ]
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 1
        txn = result[0]
        assert txn.stock_type == "ESPP"
        assert txn.acquisition_price_usd == 130.0
        assert not txn.is_long_term  # <2 years
    
    def test_skip_non_sale_transactions(self, parser):
        """Test that non-sale transactions are skipped."""
        transactions = [
            {"Action": "Deposit", "Date": "04/15/2025"},
            {"Action": "Dividend", "Date": "04/15/2025"},
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 0
    
    def test_skip_transactions_before_start_date(self, parser):
        """Test filtering by start date."""
        transactions = [
            {
                "Action": "Sale",
                "Date": "03/15/2025",  # Before start date
                "Symbol": "AAPL",
                "FeesAndCommissions": "$10.00",
                "TransactionDetails": [
                    {
                        "Details": {
                            "Type": "RS",
                            "Shares": "100",
                            "SalePrice": "$150.00",
                            "GrossProceeds": "$15000.00",
                            "VestDate": "01/15/2023",
                            "VestFairMarketValue": "$120.00"
                        }
                    }
                ]
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 0
    
    def test_proportional_fee_distribution(self, parser):
        """Test that fees are distributed proportionally across lots."""
        transactions = [
            {
                "Action": "Sale",
                "Date": "04/15/2025",
                "Symbol": "AAPL",
                "FeesAndCommissions": "$30.00",
                "TransactionDetails": [
                    {
                        "Details": {
                            "Type": "RS",
                            "Shares": "100",
                            "SalePrice": "$150.00",
                            "GrossProceeds": "$15000.00",
                            "VestDate": "01/15/2023",
                            "VestFairMarketValue": "$120.00"
                        }
                    },
                    {
                        "Details": {
                            "Type": "RS",
                            "Shares": "200",
                            "SalePrice": "$150.00",
                            "GrossProceeds": "$30000.00",
                            "VestDate": "01/15/2022",
                            "VestFairMarketValue": "$100.00"
                        }
                    }
                ]
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 2
        # First lot (100 shares) should get 1/3 of fees
        assert result[0].fees_and_commissions_usd == pytest.approx(10.0)
        # Second lot (200 shares) should get 2/3 of fees
        assert result[1].fees_and_commissions_usd == pytest.approx(20.0)


class TestSchwabIndividualParser:
    """Tests for SchwabIndividualParser class."""
    
    @pytest.fixture
    def parser(self):
        """Create parser instance."""
        return SchwabIndividualParser()
    
    def test_parse_buy_and_sell_fifo(self, parser):
        """Test FIFO matching of buy and sell transactions."""
        transactions = [
            {
                "Action": "Buy",
                "Date": "01/15/2023",
                "Symbol": "VTI",
                "Quantity": "50",
                "Price": "$200.00"
            },
            {
                "Action": "Buy",
                "Date": "06/15/2023",
                "Symbol": "VTI",
                "Quantity": "50",
                "Price": "$220.00"
            },
            {
                "Action": "Sell",
                "Date": "04/15/2025",
                "Symbol": "VTI",
                "Quantity": "30",
                "Price": "$250.00",
                "Fees & Comm": "$5.00"
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 1
        txn = result[0]
        assert txn.symbol == "VTI"
        assert txn.shares == 30
        assert txn.sale_price_usd == 250.0
        # FIFO: Should use first lot's price
        assert txn.acquisition_price_usd == 200.0
        assert txn.stock_type == "TRADE"
        assert txn.source == "Individual"
    
    def test_fifo_multiple_lots(self, parser):
        """Test selling across multiple lots."""
        transactions = [
            {
                "Action": "Buy",
                "Date": "01/15/2023",
                "Symbol": "VTI",
                "Quantity": "30",
                "Price": "$200.00"
            },
            {
                "Action": "Buy",
                "Date": "06/15/2023",
                "Symbol": "VTI",
                "Quantity": "30",
                "Price": "$220.00"
            },
            {
                "Action": "Sell",
                "Date": "04/15/2025",
                "Symbol": "VTI",
                "Quantity": "50",
                "Price": "$250.00",
                "Fees & Comm": "$10.00"
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        # Should create 2 transactions (from 2 lots)
        assert len(result) == 2
        
        # First lot (30 shares at $200)
        assert result[0].shares == 30
        assert result[0].acquisition_price_usd == 200.0
        
        # Second lot (20 shares at $220)
        assert result[1].shares == 20
        assert result[1].acquisition_price_usd == 220.0
    
    def test_reinvest_shares_treated_as_buy(self, parser):
        """Test that dividend reinvestment is treated as purchase."""
        transactions = [
            {
                "Action": "Reinvest Shares",
                "Date": "01/15/2023",
                "Symbol": "VTI",
                "Quantity": "5",
                "Price": "$200.00"
            },
            {
                "Action": "Sell",
                "Date": "04/15/2025",
                "Symbol": "VTI",
                "Quantity": "5",
                "Price": "$250.00",
                "Fees & Comm": "$1.00"
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        assert len(result) == 1
        assert result[0].acquisition_price_usd == 200.0
    
    def test_sales_before_start_date_consume_lots(self, parser):
        """Test that sales before start date consume lots for accurate FIFO."""
        transactions = [
            {
                "Action": "Buy",
                "Date": "01/15/2023",
                "Symbol": "VTI",
                "Quantity": "100",
                "Price": "$200.00"
            },
            {
                "Action": "Sell",
                "Date": "01/15/2024",  # Before start date
                "Symbol": "VTI",
                "Quantity": "50",
                "Price": "$230.00"
            },
            {
                "Action": "Sell",
                "Date": "04/15/2025",  # After start date
                "Symbol": "VTI",
                "Quantity": "30",
                "Price": "$250.00"
            }
        ]
        
        start_date = datetime(2025, 4, 1)
        result = parser.parse(transactions, start_date)
        
        # Only the second sale should be returned
        assert len(result) == 1
        assert result[0].shares == 30
        # Should use remaining lot (50 shares left after first sale)
        assert result[0].acquisition_price_usd == 200.0


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
class TestZerodhaPnLParser:
    """Tests for ZerodhaPnLParser class."""
    
    @pytest.fixture
    def parser(self):
        """Create parser instance."""
        return ZerodhaPnLParser()
    
    @pytest.fixture
    def sample_zerodha_file(self):
        """Create a sample Zerodha P&L Excel file for testing."""
        wb = Workbook()
        ws = wb.active
        
        # Row 1-9: Empty/header rows
        for i in range(1, 10):
            ws.cell(row=i, column=1, value=None)
        
        # Row 10: Statement title
        ws.cell(row=11, column=2, value="P&L Statement for Equity from 2025-04-01 to 2025-12-15")
        
        # Row 13: Summary header
        ws.cell(row=13, column=2, value="Summary")
        
        # Row 15-18: Summary values
        ws.cell(row=15, column=2, value="Charges")
        ws.cell(row=15, column=3, value=1000.50)
        
        ws.cell(row=16, column=2, value="Other Credit & Debit")
        ws.cell(row=16, column=3, value=-100.25)
        
        ws.cell(row=17, column=2, value="Realized P&L")
        ws.cell(row=17, column=3, value=50000.75)
        
        ws.cell(row=18, column=2, value="Unrealized P&L")
        ws.cell(row=18, column=3, value=-10000.00)
        
        # Row 21: Charges section
        ws.cell(row=21, column=2, value="Charges")
        
        # Row 23: Account Head header
        ws.cell(row=23, column=2, value="Account Head")
        ws.cell(row=23, column=3, value="Amount")
        
        # Row 24-32: Charges breakdown
        ws.cell(row=24, column=2, value="Brokerage - Z")
        ws.cell(row=24, column=3, value=250.50)
        
        ws.cell(row=25, column=2, value="Exchange Transaction Charges - Z")
        ws.cell(row=25, column=3, value=150.25)
        
        ws.cell(row=26, column=2, value="Securities Transaction Tax - Z")
        ws.cell(row=26, column=3, value=500.00)
        
        ws.cell(row=27, column=2, value="Stamp Duty - Z")
        ws.cell(row=27, column=3, value=100.00)
        
        # Row 38: Transaction header
        ws.cell(row=38, column=2, value="Symbol")
        ws.cell(row=38, column=3, value="ISIN")
        ws.cell(row=38, column=4, value="Quantity")
        ws.cell(row=38, column=5, value="Buy Value")
        ws.cell(row=38, column=6, value="Sell Value")
        ws.cell(row=38, column=7, value="Realized P&L")
        ws.cell(row=38, column=8, value="Realized P&L Pct.")
        ws.cell(row=38, column=9, value="Previous Closing Price")
        ws.cell(row=38, column=10, value="Open Quantity")
        ws.cell(row=38, column=11, value="Open Quantity Type")
        ws.cell(row=38, column=12, value="Open Value")
        ws.cell(row=38, column=13, value="Unrealized P&L")
        ws.cell(row=38, column=14, value="Unrealized P&L Pct.")
        
        # Row 39-41: Transaction data
        ws.cell(row=39, column=2, value="RELIANCE")
        ws.cell(row=39, column=3, value="INE002A01018")
        ws.cell(row=39, column=4, value=100)
        ws.cell(row=39, column=5, value=250000.00)
        ws.cell(row=39, column=6, value=280000.00)
        ws.cell(row=39, column=7, value=30000.00)
        ws.cell(row=39, column=8, value=12.0)
        ws.cell(row=39, column=9, value=0)
        ws.cell(row=39, column=10, value=0)
        
        ws.cell(row=40, column=2, value="TCS")
        ws.cell(row=40, column=3, value="INE467B01029")
        ws.cell(row=40, column=4, value=50)
        ws.cell(row=40, column=5, value=200000.00)
        ws.cell(row=40, column=6, value=220000.75)
        ws.cell(row=40, column=7, value=20000.75)
        ws.cell(row=40, column=8, value=10.0)
        ws.cell(row=40, column=9, value=0)
        ws.cell(row=40, column=10, value=0)
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        yield tmp_path
        
        # Cleanup
        os.unlink(tmp_path)
    
    def test_parse_realized_pnl(self, parser, sample_zerodha_file):
        """Test parsing realized P&L from Zerodha report."""
        result = parser.parse(sample_zerodha_file)
        
        assert result.source == "Zerodha Stocks"
        assert result.stcg == pytest.approx(50000.75)
        assert result.ltcg == 0.0  # Zerodha doesn't distinguish LTCG
    
    def test_parse_charges(self, parser, sample_zerodha_file):
        """Test parsing charges from Zerodha report."""
        result = parser.parse(sample_zerodha_file)
        
        assert "Brokerage" in result.charges
        assert result.charges["Brokerage"] == pytest.approx(250.50)
        assert "STT" in result.charges
        assert result.charges["STT"] == pytest.approx(500.00)
        assert "Stamp Duty" in result.charges
        assert result.charges["Stamp Duty"] == pytest.approx(100.00)
    
    def test_parse_transactions(self, parser, sample_zerodha_file):
        """Test parsing transactions from Zerodha report."""
        result = parser.parse(sample_zerodha_file)
        
        assert len(result.transactions) == 2
        
        # Check first transaction (RELIANCE)
        reliance_txn = next(t for t in result.transactions if t['symbol'] == 'RELIANCE')
        assert reliance_txn['isin'] == 'INE002A01018'
        assert reliance_txn['quantity'] == 100
        assert reliance_txn['buy_value'] == pytest.approx(250000.00)
        assert reliance_txn['sell_value'] == pytest.approx(280000.00)
        assert reliance_txn['realized_pnl'] == pytest.approx(30000.00)
        
        # Check second transaction (TCS)
        tcs_txn = next(t for t in result.transactions if t['symbol'] == 'TCS')
        assert tcs_txn['isin'] == 'INE467B01029'
        assert tcs_txn['quantity'] == 50
        assert tcs_txn['realized_pnl'] == pytest.approx(20000.75)
    
    def test_parse_negative_pnl(self, parser):
        """Test parsing negative (loss) P&L."""
        wb = Workbook()
        ws = wb.active
        
        # Minimal file with just realized P&L
        ws.cell(row=17, column=2, value="Realized P&L")
        ws.cell(row=17, column=3, value=-25000.50)
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            result = parser.parse(tmp_path)
            assert result.stcg == pytest.approx(-25000.50)
        finally:
            os.unlink(tmp_path)
    
    def test_parse_empty_file(self, parser):
        """Test parsing empty/invalid file returns empty result."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Invalid data")
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        try:
            result = parser.parse(tmp_path)
            assert result.source == "Zerodha Stocks"
            assert result.stcg == 0.0
            assert result.ltcg == 0.0
            assert len(result.transactions) == 0
        finally:
            os.unlink(tmp_path)
    
    def test_parse_nonexistent_file(self, parser):
        """Test parsing nonexistent file returns empty result."""
        result = parser.parse("/nonexistent/path/file.xlsx")
        
        assert result.source == "Zerodha Stocks"
        assert result.stcg == 0.0
        assert result.ltcg == 0.0

