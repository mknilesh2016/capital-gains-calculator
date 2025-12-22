"""
Unit tests for report generators.
"""

import pytest
import tempfile
import os
from datetime import datetime
from io import StringIO
import sys

from capital_gains.models import SaleTransaction, IndianGains, TaxData, QuarterlyData
from capital_gains.reports.console import ConsoleReporter
from capital_gains.reports.excel import ExcelReporter


class TestConsoleReporter:
    """Tests for ConsoleReporter class."""
    
    @pytest.fixture
    def reporter(self):
        """Create reporter instance."""
        return ConsoleReporter()
    
    @pytest.fixture
    def sample_transactions(self):
        """Create sample transactions."""
        return [
            SaleTransaction(
                sale_date=datetime(2025, 4, 15),
                acquisition_date=datetime(2022, 1, 15),
                stock_type="RS",
                symbol="AAPL",
                shares=100,
                sale_price_usd=150.0,
                acquisition_price_usd=120.0,
                gross_proceeds_usd=15000.0,
                source="EAC",
                sale_price_inr=12750.0,
                acquisition_price_inr=9840.0,
                capital_gain_inr=291000.0,
                capital_gain_usd=3000.0,
                holding_period_days=820,
                is_long_term=True,
                sale_exchange_rate=85.0,
                acquisition_exchange_rate=82.0,
            ),
            SaleTransaction(
                sale_date=datetime(2025, 5, 20),
                acquisition_date=datetime(2024, 6, 15),
                stock_type="TRADE",
                symbol="VTI",
                shares=50,
                sale_price_usd=250.0,
                acquisition_price_usd=220.0,
                gross_proceeds_usd=12500.0,
                source="Individual",
                sale_price_inr=21250.0,
                acquisition_price_inr=18260.0,
                capital_gain_inr=149500.0,
                capital_gain_usd=1500.0,
                holding_period_days=339,
                is_long_term=False,
                sale_exchange_rate=85.0,
                acquisition_exchange_rate=83.0,
            ),
        ]
    
    @pytest.fixture
    def sample_indian_gains(self):
        """Create sample Indian gains."""
        return [
            IndianGains(source="Indian Stocks", ltcg=50000.0, stcg=25000.0),
            IndianGains(source="Indian Mutual Funds", ltcg=30000.0, stcg=15000.0),
        ]
    
    def test_print_detailed_report(self, reporter, sample_transactions, capsys):
        """Test detailed report output."""
        reporter.print_detailed_report(sample_transactions)
        
        captured = capsys.readouterr()
        
        # Check title is present
        assert "DETAILED CAPITAL GAINS REPORT" in captured.out
        
        # Check transaction details are present
        assert "AAPL" in captured.out
        assert "VTI" in captured.out
        assert "EAC" in captured.out
        assert "Individual" in captured.out
        assert "LONG TERM" in captured.out
        assert "SHORT TERM" in captured.out
    
    def test_print_summary_report(self, reporter, sample_transactions, capsys):
        """Test summary report output."""
        reporter.print_summary_report(sample_transactions)
        
        captured = capsys.readouterr()
        
        # Check section headers
        assert "CAPITAL GAINS SUMMARY" in captured.out
        assert "TRANSACTION OVERVIEW" in captured.out
        assert "CAPITAL GAINS CLASSIFICATION" in captured.out
        
        # Check totals are shown
        assert "LONG TERM CAPITAL GAINS" in captured.out
        assert "SHORT TERM CAPITAL GAINS" in captured.out
    
    def test_print_quarterly_breakdown(self, reporter, sample_transactions, sample_indian_gains, capsys):
        """Test quarterly breakdown output."""
        result = reporter.print_quarterly_breakdown(sample_transactions, sample_indian_gains)
        
        captured = capsys.readouterr()
        
        # Check quarters are shown
        assert "Upto 15 Jun" in captured.out
        assert "16 Jun-15 Sep" in captured.out
        
        # Check result structure
        assert 'foreign' in result
        assert 'indian_stocks' in result
        assert 'indian_mf' in result
        assert 'combined' in result
    
    def test_print_grand_total(self, reporter, sample_transactions, sample_indian_gains, capsys):
        """Test grand total output."""
        reporter.print_grand_total(sample_transactions, sample_indian_gains)
        
        captured = capsys.readouterr()
        
        assert "GRAND TOTAL CAPITAL GAINS" in captured.out
        assert "Schwab" in captured.out
        assert "Indian Stocks" in captured.out
        assert "Indian Mutual Funds" in captured.out
    
    def test_empty_transactions(self, reporter, capsys):
        """Test handling empty transaction list."""
        reporter.print_detailed_report([])
        
        captured = capsys.readouterr()
        assert "DETAILED CAPITAL GAINS REPORT" in captured.out
    
    def test_single_symbol_breakdown(self, reporter, sample_transactions, capsys):
        """Test that single-symbol breakdown is skipped."""
        # Create transactions with same symbol
        same_symbol_txns = [
            SaleTransaction(
                sale_date=datetime(2025, 4, 15),
                acquisition_date=datetime(2023, 1, 15),
                stock_type="RS",
                symbol="AAPL",
                shares=100,
                sale_price_usd=150.0,
                acquisition_price_usd=120.0,
                gross_proceeds_usd=15000.0,
                capital_gain_inr=290000.0,
                is_long_term=True,
            ),
        ]
        
        reporter.print_summary_report(same_symbol_txns)
        captured = capsys.readouterr()
        
        # Should not show symbol breakdown for single symbol
        assert "BREAKDOWN BY SYMBOL" not in captured.out


class TestExcelReporter:
    """Tests for ExcelReporter class."""
    
    @pytest.fixture
    def reporter(self):
        """Create reporter instance."""
        return ExcelReporter()
    
    @pytest.fixture
    def sample_transactions(self):
        """Create sample transactions for Excel export."""
        return [
            SaleTransaction(
                sale_date=datetime(2025, 4, 15),
                acquisition_date=datetime(2022, 1, 15),
                stock_type="RS",
                symbol="AAPL",
                shares=100,
                sale_price_usd=150.0,
                acquisition_price_usd=120.0,
                gross_proceeds_usd=15000.0,
                source="EAC",
                grant_id="G12345",
                sale_price_inr=12750.0,
                acquisition_price_inr=9840.0,
                capital_gain_inr=291000.0,
                capital_gain_usd=3000.0,
                holding_period_days=820,
                is_long_term=True,
                sale_exchange_rate=85.0,
                acquisition_exchange_rate=82.0,
                fees_and_commissions_usd=10.0,
                fees_and_commissions_inr=850.0,
            ),
        ]
    
    @pytest.fixture
    def sample_indian_gains(self):
        """Create sample Indian gains for Excel export."""
        return [
            IndianGains(
                source="Indian Stocks",
                ltcg=50000.0,
                stcg=25000.0,
                transactions=[{"stock_name": "TCS", "pnl": 75000.0}],
            ),
            IndianGains(
                source="Indian Mutual Funds",
                ltcg=30000.0,
                stcg=15000.0,
                transactions=[{"scheme_name": "HDFC Equity", "stcg": 15000.0}],
            ),
        ]
    
    @pytest.fixture
    def sample_tax_data(self):
        """Create sample tax data."""
        return TaxData(
            schwab_ltcg=291000.0,
            schwab_stcg=0.0,
            indian_ltcg=80000.0,
            indian_stcg=40000.0,
            total_ltcg=371000.0,
            total_stcg=40000.0,
            ltcg_rebate=125000.0,
            rebate_used=80000.0,
            total_tax=50000.0,
            taxes_paid=20000.0,
            tax_liability=30000.0,
        )
    
    def test_export_creates_file(self, reporter, sample_transactions):
        """Test that export creates an Excel file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        try:
            result = reporter.export(
                filepath=filepath,
                transactions=sample_transactions,
            )
            
            assert result is True
            assert os.path.exists(filepath)
            assert os.path.getsize(filepath) > 0
        finally:
            if os.path.exists(filepath):
                os.unlink(filepath)
    
    def test_export_with_exchange_rates(self, reporter, sample_transactions):
        """Test export with exchange rates data."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        exchange_rates = {
            "2025-04-15": 85.0,
            "2022-01-15": 82.0,
        }
        
        try:
            result = reporter.export(
                filepath=filepath,
                transactions=sample_transactions,
                exchange_rates=exchange_rates,
            )
            
            assert result is True
            
            # Verify file can be read back
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            
            # Check sheets exist
            assert "Summary" in wb.sheetnames
            assert "Schwab Foreign Stocks" in wb.sheetnames
            assert "Exchange Rates" in wb.sheetnames
            assert "Quarterly Breakdown" in wb.sheetnames
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.unlink(filepath)
    
    def test_export_with_all_data(self, reporter, sample_transactions, 
                                   sample_indian_gains, sample_tax_data):
        """Test export with all data types."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        try:
            result = reporter.export(
                filepath=filepath,
                transactions=sample_transactions,
                exchange_rates={"2025-04-15": 85.0},
                indian_gains=sample_indian_gains,
                tax_data=sample_tax_data,
            )
            
            assert result is True
            
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            
            # Check all sheets exist
            assert "Tax Calculation" in wb.sheetnames
            # Note: Source names are mapped to display names (Indian X → Groww X)
            assert "Groww Mutual Funds" in wb.sheetnames
            assert "Groww Stocks" in wb.sheetnames
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.unlink(filepath)
    
    def test_export_empty_transactions(self, reporter):
        """Test export with empty transactions."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        try:
            result = reporter.export(
                filepath=filepath,
                transactions=[],
            )
            
            assert result is True
            assert os.path.exists(filepath)
        finally:
            if os.path.exists(filepath):
                os.unlink(filepath)
    
    def test_transaction_sheet_columns(self, reporter, sample_transactions):
        """Test that transaction sheet has correct columns."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        try:
            reporter.export(filepath=filepath, transactions=sample_transactions)
            
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb["Schwab Foreign Stocks"]
            
            # Check header row
            headers = [cell.value for cell in ws[1]]
            
            assert "S.No" in headers
            assert "Source" in headers
            assert "Sale Date" in headers
            assert "Symbol" in headers
            assert "Capital Gain (INR)" in headers
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.unlink(filepath)


class TestQuarterlyDataCalculation:
    """Tests for quarterly data calculation in ConsoleReporter."""
    
    @pytest.fixture
    def reporter(self):
        return ConsoleReporter()
    
    def test_quarterly_data_foreign(self, reporter):
        """Test foreign stocks are correctly categorized by quarter."""
        transactions = [
            SaleTransaction(
                sale_date=datetime(2025, 4, 10),  # Q1: Upto 15 Jun
                acquisition_date=datetime(2022, 1, 1),
                stock_type="RS",
                symbol="AAPL",
                shares=10,
                sale_price_usd=100.0,
                acquisition_price_usd=80.0,
                gross_proceeds_usd=1000.0,
                capital_gain_inr=50000.0,
                is_long_term=True,
            ),
            SaleTransaction(
                sale_date=datetime(2025, 7, 20),  # Q2: 16 Jun-15 Sep
                acquisition_date=datetime(2024, 6, 1),
                stock_type="TRADE",
                symbol="VTI",
                shares=5,
                sale_price_usd=200.0,
                acquisition_price_usd=180.0,
                gross_proceeds_usd=1000.0,
                capital_gain_inr=30000.0,
                is_long_term=False,
            ),
        ]
        
        # Capture output but focus on return value
        import io
        import sys
        captured = io.StringIO()
        sys.stdout = captured
        
        try:
            result = reporter.print_quarterly_breakdown(transactions)
        finally:
            sys.stdout = sys.__stdout__
        
        # Check Q1 LTCG
        assert result['foreign']["Upto 15 Jun"].ltcg == 50000.0
        assert result['foreign']["Upto 15 Jun"].stcg == 0.0
        
        # Check Q2 STCG
        assert result['foreign']["16 Jun-15 Sep"].ltcg == 0.0
        assert result['foreign']["16 Jun-15 Sep"].stcg == 30000.0


