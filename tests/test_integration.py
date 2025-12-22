"""
Integration tests for end-to-end workflows.

These tests verify that all components work together correctly
from parsing through calculation and reporting.
"""

import pytest
import json
import tempfile
import os
from datetime import datetime

from capital_gains import (
    CapitalGainsCalculator,
    TaxCalculator,
    ExchangeRateService,
    SaleTransaction,
    IndianGains,
)
from capital_gains.parsers import SchwabEACParser, SchwabIndividualParser
from capital_gains.reports import ConsoleReporter, ExcelReporter


class TestEndToEndSchwabWorkflow:
    """Integration tests for Schwab transaction processing."""
    
    @pytest.fixture
    def eac_json_data(self):
        """Sample EAC JSON data structure."""
        return {
            "Transactions": [
                {
                    "Action": "Sale",
                    "Date": "04/15/2025",
                    "Symbol": "AAPL",
                    "FeesAndCommissions": "$15.00",
                    "TransactionDetails": [
                        {
                            "Details": {
                                "Type": "RS",
                                "Shares": "100",
                                "SalePrice": "$150.00",
                                "GrossProceeds": "$15000.00",
                                "VestDate": "01/15/2022",
                                "VestFairMarketValue": "$120.00",
                                "GrantId": "RSU-2022-001"
                            }
                        },
                        {
                            "Details": {
                                "Type": "RS",
                                "Shares": "50",
                                "SalePrice": "$150.00",
                                "GrossProceeds": "$7500.00",
                                "VestDate": "07/15/2023",
                                "VestFairMarketValue": "$130.00",
                                "GrantId": "RSU-2023-001"
                            }
                        }
                    ]
                },
                {
                    "Action": "Deposit",  # Non-sale - should be ignored
                    "Date": "04/10/2025",
                    "Symbol": "CASH"
                }
            ]
        }
    
    @pytest.fixture
    def individual_json_data(self):
        """Sample Individual brokerage JSON data structure."""
        return {
            "BrokerageTransactions": [
                {
                    "Action": "Buy",
                    "Date": "01/15/2023",
                    "Symbol": "VTI",
                    "Quantity": "100",
                    "Price": "$200.00"
                },
                {
                    "Action": "Reinvest Shares",
                    "Date": "03/15/2023",
                    "Symbol": "VTI",
                    "Quantity": "5",
                    "Price": "$195.00"
                },
                {
                    "Action": "Sell",
                    "Date": "04/20/2025",
                    "Symbol": "VTI",
                    "Quantity": "50",
                    "Price": "$250.00",
                    "Fees & Comm": "$5.00"
                }
            ]
        }
    
    @pytest.fixture
    def sbi_rates(self):
        """Sample SBI rates for testing."""
        return {
            "2022-01-15": 74.5,
            "2023-01-15": 82.0,
            "2023-03-15": 82.5,
            "2023-07-15": 82.8,
            "2025-04-15": 85.0,
            "2025-04-20": 85.1,
        }
    
    def test_eac_parsing_and_calculation(self, eac_json_data, sbi_rates):
        """Test parsing EAC data and calculating gains."""
        # Setup
        parser = SchwabEACParser()
        exchange_service = ExchangeRateService()
        exchange_service.sbi_rates = sbi_rates
        calculator = CapitalGainsCalculator(exchange_rate_service=exchange_service)
        
        # Parse
        start_date = datetime(2025, 4, 1)
        transactions = parser.parse(eac_json_data["Transactions"], start_date)
        
        # Verify parsing
        assert len(transactions) == 2  # Two RSU lots from single sale
        
        # Calculate
        transactions = calculator.calculate(transactions, use_sbi=True)
        
        # Verify calculations
        for txn in transactions:
            assert txn.sale_price_inr > 0
            assert txn.acquisition_price_inr > 0
            assert txn.capital_gain_inr != 0
            assert txn.sale_exchange_rate == 85.0
        
        # First lot should be LTCG (vested 2022-01-15, sold 2025-04-15 = >2 years)
        assert transactions[0].is_long_term is True
        # Second lot should be STCG (vested 2023-07-15, sold 2025-04-15 = <2 years)
        assert transactions[1].is_long_term is False
    
    def test_individual_parsing_fifo_and_calculation(self, individual_json_data, sbi_rates):
        """Test Individual brokerage with FIFO matching."""
        # Setup
        parser = SchwabIndividualParser()
        exchange_service = ExchangeRateService()
        exchange_service.sbi_rates = sbi_rates
        calculator = CapitalGainsCalculator(exchange_rate_service=exchange_service)
        
        # Parse
        start_date = datetime(2025, 4, 1)
        transactions = parser.parse(individual_json_data["BrokerageTransactions"], start_date)
        
        # Should create sale transactions using FIFO
        assert len(transactions) == 1  # Single sale of 50 shares
        assert transactions[0].shares == 50
        assert transactions[0].acquisition_price_usd == 200.0  # FIFO uses first lot
        
        # Calculate
        transactions = calculator.calculate(transactions, use_sbi=True)
        
        # Verify
        assert transactions[0].capital_gain_inr != 0
        assert transactions[0].stock_type == "TRADE"
        assert transactions[0].source == "Individual"
    
    def test_combined_eac_and_individual(self, eac_json_data, individual_json_data, sbi_rates):
        """Test processing both EAC and Individual transactions."""
        # Setup
        eac_parser = SchwabEACParser()
        individual_parser = SchwabIndividualParser()
        exchange_service = ExchangeRateService()
        exchange_service.sbi_rates = sbi_rates
        calculator = CapitalGainsCalculator(exchange_rate_service=exchange_service)
        
        start_date = datetime(2025, 4, 1)
        
        # Parse both
        eac_txns = eac_parser.parse(eac_json_data["Transactions"], start_date)
        individual_txns = individual_parser.parse(
            individual_json_data["BrokerageTransactions"], start_date
        )
        
        # Combine
        all_transactions = eac_txns + individual_txns
        assert len(all_transactions) == 3  # 2 EAC + 1 Individual
        
        # Calculate
        all_transactions = calculator.calculate(all_transactions, use_sbi=True)
        
        # Verify all have INR values
        for txn in all_transactions:
            assert txn.capital_gain_inr != 0
            assert txn.sale_exchange_rate > 0


class TestEndToEndTaxCalculation:
    """Integration tests for tax calculation workflow."""
    
    @pytest.fixture
    def schwab_transactions(self):
        """Pre-calculated Schwab transactions."""
        return [
            SaleTransaction(
                sale_date=datetime(2025, 4, 15),
                acquisition_date=datetime(2022, 1, 15),
                stock_type="RS",
                symbol="AAPL",
                shares=100,
                sale_price_usd=150.0,
                acquisition_price_usd=120.0,
                gross_proceeds_usd=15000.0,
                capital_gain_inr=300000.0,
                is_long_term=True,
                source="EAC",
            ),
            SaleTransaction(
                sale_date=datetime(2025, 5, 20),
                acquisition_date=datetime(2024, 6, 15),
                stock_type="TRADE",
                symbol="VTI",
                shares=50,
                sale_price_usd=250.0,
                acquisition_price_usd=220.0,
                gross_proceeds_usd=12500.0,
                capital_gain_inr=150000.0,
                is_long_term=False,
                source="Individual",
            ),
        ]
    
    @pytest.fixture
    def indian_gains(self):
        """Indian stock and MF gains."""
        return [
            IndianGains(source="Indian Stocks", ltcg=200000.0, stcg=50000.0),
            IndianGains(source="Indian Mutual Funds", ltcg=100000.0, stcg=25000.0),
        ]
    
    def test_full_tax_calculation(self, schwab_transactions, indian_gains):
        """Test complete tax calculation with all sources."""
        calculator = TaxCalculator()
        
        tax_data = calculator.calculate(
            transactions=schwab_transactions,
            indian_gains=indian_gains,
            taxes_paid=50000.0
        )
        
        # Verify gains categorization
        assert tax_data.schwab_ltcg == 300000.0
        assert tax_data.schwab_stcg == 150000.0
        assert tax_data.indian_ltcg == 300000.0  # 200000 + 100000
        assert tax_data.indian_stcg == 75000.0  # 50000 + 25000
        
        # Verify totals
        assert tax_data.total_ltcg == 600000.0
        assert tax_data.total_stcg == 225000.0
        
        # Verify LTCG exemption applied to Indian gains
        assert tax_data.rebate_used == 125000.0
        assert tax_data.indian_ltcg_after_rebate == 175000.0  # 300000 - 125000
        
        # Verify taxes calculated
        assert tax_data.total_tax > 0
        assert tax_data.tax_liability == tax_data.total_tax - 50000.0
    
    def test_tax_with_losses(self):
        """Test tax calculation with capital losses."""
        transactions = [
            SaleTransaction(
                sale_date=datetime(2025, 4, 15),
                acquisition_date=datetime(2024, 6, 15),
                stock_type="TRADE",
                symbol="AMZN",
                shares=20,
                sale_price_usd=150.0,
                acquisition_price_usd=180.0,
                gross_proceeds_usd=3000.0,
                capital_gain_inr=-60000.0,  # Loss
                is_long_term=False,
            ),
        ]
        
        calculator = TaxCalculator()
        tax_data = calculator.calculate(transactions=transactions)
        
        # Loss should result in zero or minimal tax
        assert tax_data.schwab_stcg == -60000.0
        assert tax_data.foreign_stcg_taxable >= 0  # Can't be negative taxable


class TestEndToEndReporting:
    """Integration tests for report generation."""
    
    @pytest.fixture
    def complete_data(self):
        """Complete dataset for reporting."""
        transactions = [
            SaleTransaction(
                sale_date=datetime(2025, 4, 15),
                acquisition_date=datetime(2022, 1, 15),
                stock_type="RS",
                symbol="AAPL",
                shares=100,
                sale_price_usd=150.0,
                acquisition_price_usd=120.0,
                gross_proceeds_usd=15000.0,
                source="EAC",
                grant_id="RSU-2022",
                sale_price_inr=12750.0,
                acquisition_price_inr=8940.0,
                capital_gain_inr=381000.0,
                capital_gain_usd=3000.0,
                holding_period_days=820,
                is_long_term=True,
                sale_exchange_rate=85.0,
                acquisition_exchange_rate=74.5,
                fees_and_commissions_usd=10.0,
                fees_and_commissions_inr=850.0,
            ),
        ]
        
        indian_gains = [
            IndianGains(source="Indian Stocks", ltcg=50000.0, stcg=25000.0),
        ]
        
        exchange_rates = {
            "2025-04-15": 85.0,
            "2022-01-15": 74.5,
        }
        
        return transactions, indian_gains, exchange_rates
    
    def test_console_and_excel_workflow(self, complete_data, capsys):
        """Test generating both console and Excel reports."""
        transactions, indian_gains, exchange_rates = complete_data
        
        # Console reporting
        console_reporter = ConsoleReporter()
        console_reporter.print_summary_report(transactions)
        console_reporter.print_grand_total(transactions, indian_gains)
        
        captured = capsys.readouterr()
        assert "CAPITAL GAINS SUMMARY" in captured.out
        assert "GRAND TOTAL" in captured.out
        
        # Calculate tax
        tax_calculator = TaxCalculator()
        tax_data = tax_calculator.calculate(
            transactions=transactions,
            indian_gains=indian_gains,
            taxes_paid=0.0
        )
        
        # Excel reporting
        excel_reporter = ExcelReporter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        try:
            result = excel_reporter.export(
                filepath=filepath,
                transactions=transactions,
                exchange_rates=exchange_rates,
                indian_gains=indian_gains,
                tax_data=tax_data,
            )
            
            assert result is True
            assert os.path.exists(filepath)
            
            # Verify Excel structure
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            
            # All expected sheets should exist
            # Note: "Indian Stocks" is mapped to "Groww Stocks" for display
            expected_sheets = [
                "Summary", "Schwab Foreign Stocks", "Exchange Rates",
                "Quarterly Breakdown", "Groww Stocks", "Tax Calculation"
            ]
            for sheet in expected_sheets:
                assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
            
            wb.close()
        finally:
            if os.path.exists(filepath):
                os.unlink(filepath)


class TestInterfaceCompliance:
    """Tests to verify classes implement interfaces correctly."""
    
    def test_exchange_rate_service_implements_protocol(self):
        """Test ExchangeRateService implements IExchangeRateProvider."""
        from capital_gains.interfaces import IExchangeRateProvider
        
        service = ExchangeRateService()
        assert isinstance(service, IExchangeRateProvider)
    
    def test_calculator_implements_protocol(self):
        """Test CapitalGainsCalculator implements IGainsCalculator."""
        from capital_gains.interfaces import IGainsCalculator
        
        calculator = CapitalGainsCalculator()
        assert isinstance(calculator, IGainsCalculator)
    
    def test_tax_calculator_implements_protocol(self):
        """Test TaxCalculator implements ITaxCalculator."""
        from capital_gains.interfaces import ITaxCalculator
        
        calculator = TaxCalculator()
        assert isinstance(calculator, ITaxCalculator)
    
    def test_parsers_implement_protocol(self):
        """Test parsers implement ITransactionParser."""
        from capital_gains.interfaces import ITransactionParser
        
        eac_parser = SchwabEACParser()
        individual_parser = SchwabIndividualParser()
        
        assert isinstance(eac_parser, ITransactionParser)
        assert isinstance(individual_parser, ITransactionParser)


class TestFileBasedWorkflow:
    """Integration tests using actual file I/O."""
    
    def test_sbi_rates_file_loading(self):
        """Test loading SBI rates from a temp file."""
        rates_data = {
            "2025-04-01": 85.0,
            "2025-04-02": 85.1,
            "2025-04-03": 85.2,
        }
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            json.dump(rates_data, f)
            filepath = f.name
        
        try:
            service = ExchangeRateService()
            result = service.load_sbi_rates(filepath)
            
            assert result is True
            assert len(service.sbi_rates) == 3
            
            # Test rate retrieval
            rate = service.get_rate(datetime(2025, 4, 1))
            assert rate == 85.0
        finally:
            os.unlink(filepath)
    
    def test_json_transaction_file_processing(self):
        """Test processing EAC JSON file."""
        json_data = {
            "Transactions": [
                {
                    "Action": "Sale",
                    "Date": "04/15/2025",
                    "Symbol": "AAPL",
                    "FeesAndCommissions": "$10.00",
                    "TransactionDetails": [
                        {
                            "Details": {
                                "Type": "RS",
                                "Shares": "50",
                                "SalePrice": "$150.00",
                                "GrossProceeds": "$7500.00",
                                "VestDate": "01/15/2022",
                                "VestFairMarketValue": "$100.00",
                                "GrantId": "TEST-001"
                            }
                        }
                    ]
                }
            ]
        }
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            json.dump(json_data, f)
            filepath = f.name
        
        try:
            # Load and parse
            with open(filepath, 'r') as f:
                data = json.load(f)
            
            parser = SchwabEACParser()
            transactions = parser.parse(data["Transactions"], datetime(2025, 4, 1))
            
            assert len(transactions) == 1
            assert transactions[0].symbol == "AAPL"
            assert transactions[0].grant_id == "TEST-001"
        finally:
            os.unlink(filepath)


